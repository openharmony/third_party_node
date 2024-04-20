#!/usr/bin/env python
# coding=utf-8
##############################################
# Copyright (c) 2021-2022 Huawei Device Co., Ltd.
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
##############################################

import json
import os
import openpyxl


def compare_json_file(generate_json, original_json):  # 获取对比结果
    with open(generate_json, 'r', encoding='utf-8') as js1:
        data1 = json.load(js1)
    with open(original_json, 'r') as js2:
        data2 = json.load(js2)
    compare_result = []
    only_file1 = []  # 装file1独有的
    result_api = filter_compare(data1)
    for it in result_api:
        name1 = it["name"]
        key = 0
        for item in data2:
            if item["name"] == name1:
                key = 1
                compare_result.append(it)
                break
        if key == 0:
            only_file1.append(it)
    only_file2 = get_difference_data(compare_result, data2)  # 获取file2独有的
    js1.close()
    js2.close()
    return compare_result, only_file1, only_file2


def get_difference_data(compare_result, original_data):
    only_file2 = []
    for item in original_data:
        name2 = item["name"]
        key = 0
        for it in compare_result:
            name1 = it["name"]
            if name2 == name1:
                key = 1
                break
        if key == 0:
            only_file2.append(item)
    return only_file2


def filter_compare(data1):  # 获取函数和变量
    result_api = []
    for it in data1:
        get_result_api(it, result_api)
    return result_api


def get_result_api(file_data, result_api):
    if 'children' in file_data:
        for item1 in file_data["children"]:  # 抛开根节点
            if (item1["kind"] == 'FUNCTION_DECL' or item1["kind"] == 'VAR_DECL') and item1["is_extern"]:
                item = filter_func(item1)
                result_api.append(item)


def get_parm(item, parm):
    if item["parm"]:
        for i in range(len(item["parm"])):
            if item["parm"][i]["kind"] != 'PARM_DECL':
                continue
            else:
                str_parm = '{} {}'.format(item["parm"][i]["type"], item["parm"][i]["name"])
                parm.append(str_parm)
        item["parm"] = parm


def filter_func(item):
    del item["is_extern"]  # 剔除is_extern键值对，过滤后都是extern
    del item["comment"]
    if "type_ref" in list(item.keys()):
        del item["type_ref"]
    if "children" in list(item.keys()):
        del item["children"]

    item["location_path"] = item["location"]["location_path"]
    item["location"] = item["location"]["location_line"]
    if item["kind"] == 'FUNCTION_DECL':
        item["kind"] = '函数类型'
        parm = []  # 装函数参数
        if "parm" in item:
            get_parm(item, parm)
    else:
        item["kind"] = '变量类型'
        del item["is_const"]
    return item


def collated_api_data(api_data: list):
    collated_data_total = []
    for api in api_data:
        api_content = ''
        if 'node_content' in api and 'content' in api['node_content']:
            api_content = api['node_content']['content']
            api_content = api_content.replace('\r', '').replace('\n', '')
        collated_data = [
            api.get('name'),
            api_content,
            api.get('kind'),
            api.get('since'),
            api.get('syscap'),
            api.get('kit_name'),
            api.get('location_path'),
            api.get('sub_system')
        ]
        collated_data_total.append(collated_data)
    return collated_data_total


def generate_excel(array, name, generate_json_unique, original_json_unique):
    first_line_infor = ['方法名', '方法内容', '类型', '版本', 'syscap', 'kit',
                        '文件路径', '子系统']
    workbook = openpyxl.Workbook()
    work_sheet1 = workbook.active
    work_sheet1.title = '对比结果'
    work_sheet1.append(first_line_infor)
    array_update = collated_api_data(array)
    write_information_to_worksheet(work_sheet1, array_update)

    work_sheet2 = workbook.create_sheet('生成json独有')
    work_sheet2.append(first_line_infor)
    generate_json_unique_update = collated_api_data(generate_json_unique)
    write_information_to_worksheet(work_sheet2, generate_json_unique_update)

    work_sheet3 = workbook.create_sheet('原有json独有')
    write_original_infor_to_worksheet(work_sheet3, original_json_unique)
    workbook.save(name)


def write_information_to_worksheet(work_sheet, information_data):
    for data in information_data:
        write_data = data[0], data[1], data[2], data[3], data[4], \
                     data[5], data[6], data[7]
        work_sheet.append(write_data)


def write_original_infor_to_worksheet(work_sheet, original_data):
    first_line_infor = ['first_introduced', '名称']
    work_sheet.append(first_line_infor)
    collated_original_total = []
    for element in original_data:
        original_list = []
        if 'first_introduced' in element:
            original_list.append(element['first_introduced'])
        else:
            original_list.append('')
        if 'name' in element:
            original_list.append(element['name'])
        else:
            original_list.append('')
        collated_original_total.append(original_list)

    for collated_element in collated_original_total:
        write_data = collated_element[0], collated_element[1]
        work_sheet.append(write_data)


def del_repetition_value(only_file_list, compare_list):
    data = []
    for item in only_file_list:
        if item not in compare_list:
            data.append(item)
    return data


def get_json_file(json_file_new, json_file):  # 获取生成的json文件
    json_file1 = r'{}'.format(json_file_new)  # 获取要对比的json文件
    json_file2 = json_file
    head_name = os.path.splitext(json_file1)  # 去掉文件名后缀
    head_name = head_name[0] + '.xlsx'  # 加后缀
    result_list = []
    only_file1 = []
    only_file2 = []
    for item in json_file2:  # 对比每一个json(目录下的)
        # 对比两个json文件
        result_list_part, only_file1_part, only_file2_part = compare_json_file(json_file1, item)
        result_list.extend(result_list_part)
        only_file1.extend(only_file1_part)
        only_file2.extend(only_file2_part)
    only_file1_new = del_repetition_value(only_file1, result_list)
    return result_list, head_name, only_file1_new, only_file2  # 返回对比数据，和所需表格名


def get_api_data(parser_data, excel_file_name):
    generate_json_unique = []
    original_json_unique = []
    api_data_list = filter_compare(parser_data)
    generate_excel(api_data_list, excel_file_name, generate_json_unique, original_json_unique)
